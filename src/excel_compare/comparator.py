"""Excel comparator core logic."""

from dataclasses import dataclass
from typing import List, Dict, Optional, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class CellDiff:
    """Represents a difference in a single cell."""
    sheet_name: str
    row: int
    column: str
    old_value: Any
    new_value: Any


@dataclass
class RowDiff:
    """Represents a row difference."""
    sheet_name: str
    row_index: int
    diff_type: str  # 'added', 'deleted', 'modified'
    old_data: Optional[Dict[str, Any]] = None
    new_data: Optional[Dict[str, Any]] = None
    cell_diffs: List[CellDiff] = None


@dataclass
class ColumnDiff:
    """Represents a column structure difference."""
    sheet_name: str
    column_name: str
    diff_type: str  # 'added', 'deleted'


@dataclass
class SheetDiff:
    """Represents differences in a sheet."""
    sheet_name: str
    diff_type: str  # 'added', 'deleted', 'modified'
    column_diffs: List[ColumnDiff] = None
    row_diffs: List[RowDiff] = None


@dataclass
class CompareResult:
    """Result of comparing two Excel files."""
    file1: str
    file2: str
    sheet_diffs: List[SheetDiff]
    summary: Dict[str, int]


class ExcelComparator:
    """Compares two Excel files and identifies differences."""

    def __init__(self, file1_path: str, file2_path: str):
        """
        Initialize the comparator with two file paths.

        Args:
            file1_path: Path to the first Excel file
            file2_path: Path to the second Excel file
        """
        self.file1_path = file1_path
        self.file2_path = file2_path
        # 使用纯openpyxl：{sheet_name: {(row, col): value}}
        self.data1_dict = {}
        self.data2_dict = {}
        self.sheet_names_1 = []
        self.sheet_names_2 = []

    def load_files(self) -> None:
        """Load both Excel files using pure openpyxl - returns dictionary of cell values."""

        def read_sheet_data(wb, sheet_name):
            """使用openpyxl读取sheet的所有数据，只存储非空单元格"""
            sheet = wb[sheet_name]
            data_dict = {}

            # 动态遍历所有有数据的单元格，不依赖max_column
            for row_cells in sheet.iter_rows():
                for cell in row_cells:
                    if cell.value is not None:
                        col_letter = get_column_letter(cell.column)
                        data_dict[(cell.row, col_letter)] = cell.value

            return data_dict if data_dict else None

        # Load file1
        wb1 = load_workbook(self.file1_path, read_only=True, data_only=True)
        self.sheet_names_1 = wb1.sheetnames
        for sheet in self.sheet_names_1:
            data = read_sheet_data(wb1, sheet)
            if data:
                self.data1_dict[sheet] = data
        wb1.close()

        # Load file2
        wb2 = load_workbook(self.file2_path, read_only=True, data_only=True)
        self.sheet_names_2 = wb2.sheetnames
        for sheet in self.sheet_names_2:
            data = read_sheet_data(wb2, sheet)
            if data:
                self.data2_dict[sheet] = data
        wb2.close()

    def compare(self) -> CompareResult:
        """
        Compare the two Excel files.

        Returns:
            CompareResult object containing all differences
        """
        self.load_files()

        sheet_diffs = []

        # Get all unique sheet names
        all_sheets = set(self.sheet_names_1) | set(self.sheet_names_2)

        for sheet_name in all_sheets:
            if sheet_name in self.sheet_names_1 and sheet_name not in self.sheet_names_2:
                sheet_diffs.append(SheetDiff(
                    sheet_name=sheet_name,
                    diff_type='deleted'
                ))
            elif sheet_name not in self.sheet_names_1 and sheet_name in self.sheet_names_2:
                sheet_diffs.append(SheetDiff(
                    sheet_name=sheet_name,
                    diff_type='added'
                ))
            else:
                # Both files have this sheet, compare content
                diff = self._compare_sheet(sheet_name)
                if diff:
                    sheet_diffs.append(diff)

        # Generate summary
        summary = self._generate_summary(sheet_diffs)

        return CompareResult(
            file1=self.file1_path,
            file2=self.file2_path,
            sheet_diffs=sheet_diffs,
            summary=summary
        )

    def _is_empty_row(self, data_dict: Dict, row: int) -> bool:
        """判断是否为空行（没有任何非空数据）"""
        for (r, col), value in data_dict.items():
            if r == row and value is not None and str(value).strip() != '':
                return False
        return True

    def _compare_sheet(self, sheet_name: str) -> Optional[SheetDiff]:
        """
        Compare a single sheet between both files.

        Args:
            sheet_name: Name of the sheet to compare

        Returns:
            SheetDiff object if differences found, None otherwise
        """
        data1 = self.data1_dict[sheet_name]
        data2 = self.data2_dict[sheet_name]

        column_diffs = []
        row_diffs = []

        # 获取所有行号和列号
        rows1 = {row for row, _ in data1.keys()}
        rows2 = {row for row, _ in data2.keys()}
        cols1 = {col for _, col in data1.keys()}
        cols2 = {col for _, col in data2.keys()}

        max_row1 = max(rows1) if rows1 else 0
        max_row2 = max(rows2) if rows2 else 0

        # 计算最大列号（将列字母转换为数字）
        def col_letter_to_num(col_letter):
            """将列字母转换为数字（A=1, B=2, ...）"""
            num = 0
            for char in col_letter:
                num = num * 26 + (ord(char.upper()) - ord('A') + 1)
            return num

        max_col1 = max([col_letter_to_num(col) for _, col in data1.keys()]) if data1 else 0
        max_col2 = max([col_letter_to_num(col) for _, col in data2.keys()]) if data2 else 0
        max_col = max(max_col1, max_col2)

        # Compare columns
        all_cols = cols1 | cols2
        for col in cols1 - cols2:
            column_diffs.append(ColumnDiff(
                sheet_name=sheet_name,
                column_name=str(col),
                diff_type='deleted'
            ))

        for col in cols2 - cols1:
            column_diffs.append(ColumnDiff(
                sheet_name=sheet_name,
                column_name=str(col),
                diff_type='added'
            ))

        # Only compare rows if there are common columns and C列 exists
        common_cols = cols1 & cols2
        all_compare_cols = cols1 | cols2  # 所有列，用于完整对比
        if common_cols and 'C' in common_cols:
            # 使用C列（场所名称）作为唯一标识符进行匹配
            # 关键：通过场所名称匹配，而不是按行号，这样可以避免"增加行导致错位"的问题

            # 收集file1中所有有效行（有场所名称的行）
            rows_by_name1 = {}  # {场所名称: (行号, 行数据字典)}
            for row in range(1, max_row1 + 1):
                # 跳过空行
                if self._is_empty_row(data1, row):
                    continue
                site_name = data1.get((row, 'C'))
                if site_name is not None and str(site_name).strip() != '':
                    # 收集该行的所有列数据（包括两个文件的所有列）
                    row_data = {}
                    for col in all_compare_cols:
                        row_data[col] = data1.get((row, col))
                    rows_by_name1[str(site_name).strip()] = (row, row_data)

            # 收集file2中所有有效行（有场所名称的行）
            rows_by_name2 = {}  # {场所名称: (行号, 行数据字典)}
            for row in range(1, max_row2 + 1):
                # 跳过空行
                if self._is_empty_row(data2, row):
                    continue
                site_name = data2.get((row, 'C'))
                if site_name is not None and str(site_name).strip() != '':
                    # 收集该行的所有列数据（包括两个文件的所有列）
                    row_data = {}
                    for col in all_compare_cols:
                        row_data[col] = data2.get((row, col))
                    rows_by_name2[str(site_name).strip()] = (row, row_data)

            # 获取所有场所名称
            all_names = set(rows_by_name1.keys()) | set(rows_by_name2.keys())

            # 值标准化函数
            def normalize_value(v):
                """标准化单元格值"""
                if v is None:
                    return None
                s = str(v).strip()
                return None if s == '' else s

            # 按场所名称排序遍历
            for site_name in sorted(all_names):
                in_file1 = site_name in rows_by_name1
                in_file2 = site_name in rows_by_name2

                if not in_file1 and in_file2:
                    # 新增的行 - 使用file2的行号
                    excel_row_num, row_data = rows_by_name2[site_name]
                    row_diffs.append(RowDiff(
                        sheet_name=sheet_name,
                        row_index=excel_row_num,
                        diff_type='added',
                        new_data=row_data
                    ))

                elif in_file1 and not in_file2:
                    # 删除的行 - 使用file1的行号
                    excel_row_num, row_data = rows_by_name1[site_name]
                    row_diffs.append(RowDiff(
                        sheet_name=sheet_name,
                        row_index=excel_row_num,
                        diff_type='deleted',
                        old_data=row_data
                    ))

                elif in_file1 and in_file2:
                    # 两个文件都有此场所，比较内容
                    excel_row_num1, row_data1 = rows_by_name1[site_name]
                    excel_row_num2, row_data2 = rows_by_name2[site_name]

                    # 对比所有列的单元格差异
                    cell_diffs = []
                    for col in all_compare_cols:
                        # 跳过序号列（A列）和场所名称列（C列）
                        if col == 'A' or col == 'C':
                            continue

                        val1 = row_data1.get(col)
                        val2 = row_data2.get(col)

                        norm_val1 = normalize_value(val1)
                        norm_val2 = normalize_value(val2)

                        if norm_val1 != norm_val2:
                            cell_diffs.append(CellDiff(
                                sheet_name=sheet_name,
                                row=excel_row_num2,  # 使用file2的行号
                                column=str(col),
                                old_value=norm_val1,
                                new_value=norm_val2
                            ))

                    if cell_diffs:
                        row_diffs.append(RowDiff(
                            sheet_name=sheet_name,
                            row_index=excel_row_num2,  # 使用file2的行号
                            diff_type='modified',
                            old_data=row_data1,
                            new_data=row_data2,
                            cell_diffs=cell_diffs
                        ))

            # 按行号排序，使显示更有序
            row_diffs.sort(key=lambda x: x.row_index)

        # Only return SheetDiff if there are differences
        if column_diffs or row_diffs:
            return SheetDiff(
                sheet_name=sheet_name,
                diff_type='modified',
                column_diffs=column_diffs if column_diffs else None,
                row_diffs=row_diffs if row_diffs else None
            )

        return None

    def _generate_summary(self, sheet_diffs: List[SheetDiff]) -> Dict[str, int]:
        """
        Generate a summary of all differences.

        Args:
            sheet_diffs: List of sheet differences

        Returns:
            Dictionary containing summary statistics
        """
        summary = {
            'total_sheets': 0,
            'added_sheets': 0,
            'deleted_sheets': 0,
            'modified_sheets': 0,
            'added_columns': 0,
            'deleted_columns': 0,
            'added_rows': 0,
            'deleted_rows': 0,
            'modified_cells': 0,
            'total_differences': 0
        }

        for diff in sheet_diffs:
            summary['total_sheets'] += 1
            summary['total_differences'] += 1

            if diff.diff_type == 'added':
                summary['added_sheets'] += 1
            elif diff.diff_type == 'deleted':
                summary['deleted_sheets'] += 1
            elif diff.diff_type == 'modified':
                summary['modified_sheets'] += 1

                if diff.column_diffs:
                    for col_diff in diff.column_diffs:
                        if col_diff.diff_type == 'added':
                            summary['added_columns'] += 1
                        elif col_diff.diff_type == 'deleted':
                            summary['deleted_columns'] += 1
                        summary['total_differences'] += 1

                if diff.row_diffs:
                    for row_diff in diff.row_diffs:
                        if row_diff.diff_type == 'added':
                            summary['added_rows'] += 1
                        elif row_diff.diff_type == 'deleted':
                            summary['deleted_rows'] += 1
                        elif row_diff.diff_type == 'modified':
                            if row_diff.cell_diffs:
                                summary['modified_cells'] += len(row_diff.cell_diffs)
                        summary['total_differences'] += 1

        return summary
